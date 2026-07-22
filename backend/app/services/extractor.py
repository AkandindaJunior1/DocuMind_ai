"""
Text extraction service.
Takes a file path and returns the raw text content, page by page.
Supports: PDF, DOCX, TXT, CSV, XLSX
"""
import os
from PyPDF2 import PdfReader
from docx import Document as DocxDocument
import csv
import openpyxl

def extract_text(file_path: str, file_type: str) -> list[dict]:
    """
    Returns a list of dicts: [{"page": 1, "text": "..."}, ...]
    For formats without real pages (txt, csv), we return a single "page".
    """
    file_type = file_type.lower()

    if file_type == "pdf":
        return _extract_pdf(file_path)
    elif file_type == "docx":
        return _extract_docx(file_path)
    elif file_type == "txt":
        return _extract_txt(file_path)
    elif file_type == "csv":
        return _extract_csv(file_path)
    elif file_type == "xlsx":
        return _extract_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file type: {file_type}")


def _extract_pdf(file_path: str) -> list[dict]:
    reader = PdfReader(file_path)
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append({"page": i + 1, "text": text.strip()})
    return pages


def _extract_docx(file_path: str) -> list[dict]:
    doc = DocxDocument(file_path)
    full_text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
    if not full_text.strip():
        return []
    return [{"page": 1, "text": full_text}]


def _extract_txt(file_path: str) -> list[dict]:
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()
    if not text.strip():
        return []
    return [{"page": 1, "text": text.strip()}]


def _extract_csv(file_path: str) -> list[dict]:
    rows = []
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(", ".join(row))
    text = "\n".join(rows)
    if not text.strip():
        return []
    return [{"page": 1, "text": text}]


def _extract_xlsx(file_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    pages = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            row_text = ", ".join([str(cell) if cell is not None else "" for cell in row])
            rows.append(row_text)
        text = "\n".join(rows)
        if text.strip():
            pages.append({"page": 1, "text": f"Sheet: {sheet_name}\n{text}"})
    wb.close()
    return pages
